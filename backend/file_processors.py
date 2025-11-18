"""
File processors for ZIP, PDF, Excel, Shapefile, and other formats
"""
import zipfile
import io
import os
from typing import List, Dict, Any, BinaryIO
from PyPDF2 import PdfReader
import openpyxl
import fiona
import json
from shapely.geometry import shape, mapping
import csv


class FileProcessor:
    """Base class for file processors"""

    @staticmethod
    def process_zip(file_content: bytes, batch_id: str) -> Dict[str, Any]:
        """
        Extract and process ZIP file containing mixed data types

        Returns:
            {
                'csvs': [list of CSV data dicts],
                'pdfs': [list of PDF data dicts],
                'shapefiles': [list of shapefile data dicts],
                'excels': [list of Excel data dicts],
                'geojsons': [list of GeoJSON data dicts]
            }
        """
        results = {
            'csvs': [],
            'pdfs': [],
            'shapefiles': [],
            'excels': [],
            'geojsons': [],
            'errors': []
        }

        try:
            with zipfile.ZipFile(io.BytesIO(file_content)) as zip_file:
                file_list = zip_file.namelist()
                print(f"📦 ZIP contains {len(file_list)} files")

                for filename in file_list:
                    # Skip directories and hidden files
                    if filename.endswith('/') or filename.startswith('__MACOSX') or filename.startswith('.'):
                        continue

                    try:
                        file_data = zip_file.read(filename)
                        file_ext = filename.lower().split('.')[-1]

                        if file_ext == 'csv':
                            csv_data = FileProcessor.process_csv(file_data, filename, batch_id)
                            results['csvs'].append(csv_data)

                        elif file_ext == 'pdf':
                            pdf_data = FileProcessor.process_pdf(file_data, filename, batch_id)
                            results['pdfs'].append(pdf_data)

                        elif file_ext in ['xls', 'xlsx']:
                            excel_data = FileProcessor.process_excel(file_data, filename, batch_id)
                            results['excels'].append(excel_data)

                        elif file_ext in ['geojson', 'json']:
                            geojson_data = FileProcessor.process_geojson(file_data, filename, batch_id)
                            results['geojsons'].append(geojson_data)

                        elif file_ext == 'shp':
                            # Shapefiles need additional files (.shx, .dbf, .prj)
                            shp_data = FileProcessor.process_shapefile_from_zip(zip_file, filename, batch_id)
                            if shp_data:
                                results['shapefiles'].append(shp_data)

                    except Exception as e:
                        print(f"⚠️ Error processing {filename}: {e}")
                        results['errors'].append({
                            'filename': filename,
                            'error': str(e)
                        })

        except zipfile.BadZipFile as e:
            print(f"❌ Invalid ZIP file: {e}")
            results['errors'].append({'error': 'Invalid ZIP file format'})

        return results

    @staticmethod
    def process_csv(file_content: bytes, filename: str, batch_id: str) -> Dict[str, Any]:
        """Process CSV file and extract data"""
        try:
            csv_text = file_content.decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(csv_text))
            rows = list(csv_reader)

            return {
                'filename': filename,
                'batch_id': batch_id,
                'row_count': len(rows),
                'columns': csv_reader.fieldnames if csv_reader.fieldnames else [],
                'data': rows
            }
        except Exception as e:
            print(f"❌ CSV processing error for {filename}: {e}")
            return {'filename': filename, 'error': str(e)}

    @staticmethod
    def process_pdf(file_content: bytes, filename: str, batch_id: str) -> Dict[str, Any]:
        """Extract text from PDF file"""
        try:
            pdf_reader = PdfReader(io.BytesIO(file_content))
            page_count = len(pdf_reader.pages)

            # Extract text from all pages
            full_text = ""
            for page_num, page in enumerate(pdf_reader.pages):
                page_text = page.extract_text()
                full_text += f"\n--- Page {page_num + 1} ---\n{page_text}"

            # Extract metadata
            metadata = {}
            if pdf_reader.metadata:
                metadata = {
                    'title': pdf_reader.metadata.get('/Title', ''),
                    'author': pdf_reader.metadata.get('/Author', ''),
                    'subject': pdf_reader.metadata.get('/Subject', ''),
                    'creator': pdf_reader.metadata.get('/Creator', ''),
                }

            return {
                'filename': filename,
                'batch_id': batch_id,
                'page_count': page_count,
                'text_content': full_text,
                'metadata': metadata,
                'file_size': len(file_content)
            }
        except Exception as e:
            print(f"❌ PDF processing error for {filename}: {e}")
            return {'filename': filename, 'error': str(e)}

    @staticmethod
    def process_excel(file_content: bytes, filename: str, batch_id: str) -> Dict[str, Any]:
        """Process Excel file and extract data from all sheets"""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            sheets_data = {}

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Get all values as list of lists
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    rows.append(list(row))

                sheets_data[sheet_name] = rows

            return {
                'filename': filename,
                'batch_id': batch_id,
                'sheet_count': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames,
                'data': sheets_data
            }
        except Exception as e:
            print(f"❌ Excel processing error for {filename}: {e}")
            return {'filename': filename, 'error': str(e)}

    @staticmethod
    def process_geojson(file_content: bytes, filename: str, batch_id: str) -> Dict[str, Any]:
        """Process GeoJSON file"""
        try:
            geojson_data = json.loads(file_content.decode('utf-8'))

            return {
                'filename': filename,
                'batch_id': batch_id,
                'feature_count': len(geojson_data.get('features', [])),
                'geojson': geojson_data
            }
        except Exception as e:
            print(f"❌ GeoJSON processing error for {filename}: {e}")
            return {'filename': filename, 'error': str(e)}

    @staticmethod
    def process_shapefile_from_zip(zip_file: zipfile.ZipFile, shp_filename: str, batch_id: str) -> Dict[str, Any]:
        """
        Process shapefile from ZIP (requires .shp, .shx, .dbf files)
        Converts to GeoJSON format
        """
        try:
            # Extract base name without extension
            base_name = os.path.splitext(shp_filename)[0]

            # Required files for shapefile
            required_extensions = ['.shp', '.shx', '.dbf']
            optional_extensions = ['.prj', '.cpg']

            # Create temporary directory for shapefile components
            temp_files = {}

            for ext in required_extensions + optional_extensions:
                file_path = f"{base_name}{ext}"
                if file_path in zip_file.namelist():
                    temp_files[ext] = zip_file.read(file_path)

            # Check if we have all required files
            if not all(ext in temp_files for ext in required_extensions):
                print(f"⚠️ Shapefile {shp_filename} missing required files")
                return None

            # Write temporary files to disk (fiona requires files on disk)
            import tempfile
            temp_dir = tempfile.mkdtemp()

            try:
                for ext, content in temp_files.items():
                    with open(os.path.join(temp_dir, f"{os.path.basename(base_name)}{ext}"), 'wb') as f:
                        f.write(content)

                # Read shapefile with fiona
                shp_path = os.path.join(temp_dir, f"{os.path.basename(base_name)}.shp")

                with fiona.open(shp_path) as src:
                    # Convert to GeoJSON
                    features = []
                    for feature in src:
                        features.append({
                            'type': 'Feature',
                            'geometry': feature['geometry'],
                            'properties': feature['properties']
                        })

                    geojson_data = {
                        'type': 'FeatureCollection',
                        'features': features
                    }

                    return {
                        'filename': shp_filename,
                        'batch_id': batch_id,
                        'feature_count': len(features),
                        'crs': src.crs.to_dict() if src.crs else None,
                        'geojson': geojson_data
                    }

            finally:
                # Cleanup temporary files
                import shutil
                shutil.rmtree(temp_dir)

        except Exception as e:
            print(f"❌ Shapefile processing error for {shp_filename}: {e}")
            return None


class DocumentChunker:
    """Split documents into chunks for RAG embedding"""

    @staticmethod
    def chunk_text(text: str, chunk_size: int = 1000, overlap: int = 200) -> List[str]:
        """
        Split text into overlapping chunks

        Args:
            text: Full text to chunk
            chunk_size: Target size of each chunk in characters
            overlap: Number of characters to overlap between chunks

        Returns:
            List of text chunks
        """
        if not text or len(text) == 0:
            return []

        chunks = []
        start = 0

        while start < len(text):
            end = start + chunk_size

            # Try to break at sentence boundary
            if end < len(text):
                # Look for sentence endings
                sentence_end = max(
                    text.rfind('. ', start, end),
                    text.rfind('! ', start, end),
                    text.rfind('? ', start, end),
                    text.rfind('\n\n', start, end)
                )

                if sentence_end > start:
                    end = sentence_end + 1

            chunk = text[start:end].strip()
            if chunk:
                chunks.append(chunk)

            start = end - overlap if end < len(text) else end

        return chunks

    @staticmethod
    def get_chunk_metadata(chunk: str) -> Dict[str, int]:
        """Get metadata about a text chunk"""
        return {
            'char_count': len(chunk),
            'word_count': len(chunk.split())
        }
